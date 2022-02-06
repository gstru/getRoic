from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
import pandas as pd
import numpy as np
import yfinance as yf
import statistics as st
import openpyxl
import requests
import time
import pyfiglet
from rich.console import Console
from rich import print
import difflib
import math

# Rich Console
console = Console()
# Carico Carico File Excel Finviz Industry
filename = 'PATH/finviz.xlsx'
finviz = pd.read_excel(filename)
# Carico File Excel Watchlist
filename = 'PATH/WatchList.xlsx'
wb = openpyxl.load_workbook(filename)
ws = wb['YNAME']
stocks = []
column = ws['A']
for x in range(0, len(column)):
    stocks.append(column[x].value)
print('Ho caricato le lista Stocks')
print("")

# Carico Aziende già analizzate
ws = wb['MULTIPLI']
stocks_multipli = []
column = ws['A']
for x in range(1, len(column)):
    stocks_multipli.append(column[x].value)
print('Ho caricato le lista delle Stocks già analizzate')

# Carico Performance Indice S&P500 ultimi 10 anni
table_sp500 = pd.read_html('https://www.macrotrends.net/2526/sp-500-historical-annual-returns')
table_sp500 = table_sp500[0]
table_sp500 = table_sp500.values
annual_change = []
for t in table_sp500:
    p = t[6]
    annual_change.append(p)
annual_change = [element.replace('%', '') for element in annual_change]
sp500 = np.flip(annual_change)
sp500 = sp500[-10:]
sp500 = sp500.astype(np.float)
sp500 = [element / 100 for element in sp500]
print("Carico Performance Indice S&P500 ultimi 10 anni")
# Calcolo il Rendimento di Mercato
rdm = st.mean(sp500)
# Rendimento Free Risk -> Tasso di Interesse a 10 anni
table_wgb = pd.read_html('http://www.worldgovernmentbonds.com/')
table_wgb = table_wgb[1]
wgb_country = table_wgb["Country"].values
# Tasso di crescita PIL - GDP - Growth Rate
gdp = [2.56,1.55,2.25,1.84,2.53,3.08,1.71,2.33,3.00,2.16,-3.49]
gdp = [element / 100 for element in gdp]
gdp = st.mean(gdp)



# Carico Settings Selenium
options = webdriver.ChromeOptions()
options.add_extension("PATH/ublock.crx")
options.add_extension("PATH/cookie.crx")
options.add_argument('--disable-gpu')
options.add_argument("--log-level=0")
options.add_argument("--log-level=1")
options.add_argument("--log-level=2")
options.add_argument("--log-level=3")
options.add_argument("--ignore-certificate-errors")
options.add_argument("--start-maximized")
driver = webdriver.Chrome(executable_path=r'PATH/chromedriver.exe', options=options)

for s in stocks:
    try:
        # Set Verify
        verify = []

        # Yahoo Finance Stock
        stock = yf.Ticker(s)
        info = stock.info
        financials = stock.financials
        balance = stock.balance_sheet
        cashflow = stock.cashflow
        # Nome Stock
        name = info['shortName']
        if name in stocks_multipli:
            continue
        else:
            verify.append(name)
        # Country
        country = info['country']
        # Prezzo Stock
        price = info['regularMarketPrice']
        # Sector
        sector = info['sector']
        # Industry
        industry = info['industry']
        # Beta Stock
        beta = info['beta']
        # MarketCap
        market_cap = info['marketCap']
        # N. Shares
        n_share = info['sharesOutstanding']
        # Finviz Settings
        industries = finviz["Name"].values
        match_industries = difflib.get_close_matches(industry, industries)
        match_industries = match_industries[0]
        finviz_indice = finviz[finviz["Name"]==match_industries].index.values
        finviz_industry = finviz.iloc[finviz_indice].to_numpy()
        # Rendimento Free Risk -> Tasso di Interesse a 10 anni
        match_country = difflib.get_close_matches(country, wgb_country)
        match_result_wgb = match_country[0]
        index_wgb = table_wgb[table_wgb["Country"]==match_result_wgb].index.values
        row_wgb = table_wgb.iloc[index_wgb].to_numpy()
        rfr = row_wgb[0,3]
        rfr = rfr.replace('%', '')
        rfr = float(rfr)
        rfr = rfr / 100
#-----------------------------------------------------#
        # Price / Earnings
#-----------------------------------------------------#
        try:
            trailingPE = info['trailingPE']
        except:
            trailingPE = info['forwardPE']
        # Price / Earnings Industry
        trailingPE_industry = finviz_industry[0,2]
        # Check Price / Earnings --> 1° Multiplo
        if trailingPE < trailingPE_industry:
            verify.append("buono")
        elif trailingPE == trailingPE_industry:
            verify.append("così così")
        else:
            verify.append("non ci siamo")
#-----------------------------------------------------#
        # PEG Ratio
#-----------------------------------------------------#

        peg_ratio = info['pegRatio']
        if (peg_ratio is None):
            url = 'https://finance.yahoo.com/quote/' + s + '/financials'

            driver.get(url)
            time.sleep(10)
            righe = []
            indice = []

            tabella = driver.find_element(By.CLASS_NAME, "W\(100\%\).Whs\(nw\).Ovx\(a\).BdT.Bdtc\(\$seperatorColor\)")

            titoli_tabella_1 = tabella.find_element(By.CLASS_NAME, "D\(tbhg\)")
            titoli_span = titoli_tabella_1.find_elements(By.TAG_NAME, "span")
            for m in titoli_span:
                indice.append(m.text)

            indici_tabella_1 = tabella.find_elements(By.CLASS_NAME, "D\(tbr\).fi-row.Bgc\(\$hoverBgColor\)\:h")
            for i in indici_tabella_1:
                indici_tabella_2 = i.find_elements(By.TAG_NAME, "div")
                array = []
                for j in indici_tabella_2:
                    span = j.find_elements(By.TAG_NAME, "span")
                    for k in span:
                        number = (k.text).replace(",", "")
                        array.append(number)
                array = np.delete(array, 0)
                righe.append(array)

            tabelle = pd.DataFrame(righe, columns = indice)
            eps_indice = tabelle[tabelle["Breakdown"]=="Basic EPS"].index.values
            basic_eps = tabelle.iloc[eps_indice].to_numpy()
            # Basic EPS
            basic_eps = np.delete(basic_eps, 0)
            basic_eps = basic_eps.astype(np.float)
            lenght_eps = len(basic_eps)
            if lenght_eps > 4:
                basic_eps = np.delete(basic_eps, 0)
            basic_eps = np.flip(basic_eps)
            tassi_crescita_eps = []
            for i in range(0,len(basic_eps)):

                if (i > 0):
                    # Calcolo i Tassi di Crescita
                    element = basic_eps[i] / basic_eps[i - 1]
                    element = (element - 1)*100
                    element = round(element,3)
                    tassi_crescita_eps.append(element)
                    element = None

                else:
                    continue

            media_tassi_crescita_eps = st.mean(tassi_crescita_eps)
            peg_ratio = trailingPE / media_tassi_crescita_eps


        # Check PEG Ratio --> 2° Multiplo
        if peg_ratio < 1:
            verify.append("buono")
        else:
            verify.append("non ci siamo")
#-----------------------------------------------------#
        # Price / Cash Flow (P/CF)
#-----------------------------------------------------#
        operating_cashflow = cashflow.loc['Total Cash From Operating Activities'].to_numpy()
        operating_cashflow = np.flip(operating_cashflow)
        operating_cashflow = operating_cashflow[-1]
        price_cf = price * (n_share / operating_cashflow)
        # Check Price / Cash Flow (P/CF) --> 3° Multiplo
        if price_cf < trailingPE:
            verify.append("buono")
        elif price_cf <= (2 * trailingPE):
            verify.append("così così")
        else:
            verify.append("non ci siamo")
#-----------------------------------------------------#
        # Price / Book Value (P/B)
#-----------------------------------------------------#
        price_to_book = info['priceToBook']
        price_to_book_settore = finviz_industry[0,6]
        # Check Price / Book Value (P/B) --> 4° Multiplo
        try:
            if price_to_book < price_to_book_settore:
                verify.append("buono")
            else:
                verify.append("non ci siamo")
        except:
            verify.append("errore")
#-----------------------------------------------------#
        # Enterprise Value / EBITDA (EV/EBITDA)
#-----------------------------------------------------#
        enterprise_to_ebitda = info['enterpriseToEbitda']
        table_siblis = pd.read_html('https://siblisresearch.com/data/ev-ebitda-multiple/')
        table_siblis = table_siblis[0]
        GICS_Sector = table_siblis["GICS Sector"].values
        match_sectors = difflib.get_close_matches(sector, GICS_Sector)
        match_result_siblis = match_sectors[0]
        index_siblis = table_siblis[table_siblis["GICS Sector"]==match_result_siblis].index.values
        row_siblis = table_siblis.iloc[index_siblis].to_numpy()
        enterprise_to_ebitda_settore = row_siblis[0,1]
        # Check EV/EBITDA --> 5° Multiplo
        try:
            if enterprise_to_ebitda < enterprise_to_ebitda_settore:
                verify.append("buono")
            else:
                verify.append("non ci siamo")
        except:
            verify.append("errore")
#-----------------------------------------------------#
        # MarketCap
#-----------------------------------------------------#
        # Check MarketCap --> 6° Multiplo
        if country == "United States":
            if market_cap > 2000000:
                verify.append("buono")
            else:
                verify.append("non ci siamo")
        else:
            if market_cap > 500000:
                verify.append("buono")
            else:
                verify.append("non ci siamo")
#-----------------------------------------------------#
        # Quick Ratio
#-----------------------------------------------------#
        quick_ratio = info['quickRatio']
        # Check Quick Ratio --> 7° Multiplo
        if quick_ratio > 1.5 :
            verify.append("buono")
        else:
            verify.append("non ci siamo")
#-----------------------------------------------------#
        # Currents Assets / Long Term Debt > 0.90
#-----------------------------------------------------#
        # Currents Assets
        current_assets = balance.loc['Total Current Assets'].to_numpy()
        current_assets = np.flip(current_assets)
        current_assets = current_assets[-1]
        # Long Term Debt
        try:
            long_term_debt = balance.loc['Long Term Debt'].to_numpy()
            # Long Term Debt - Flippato
            long_term_debt = np.flip(long_term_debt)
            long_term_debt_1 = long_term_debt[-1]
            check = math.isnan(long_term_debt_1)
            if check == True:
                long_term_debt = long_term_debt[-2]
            else:
                long_term_debt = long_term_debt_1
        except:
            long_term_debt = 0
        # Currents Assets / Long Term Debt --> 8° Multiplo
        if (current_assets / long_term_debt) > 0.9 :
            verify.append("buono")
        else:
            verify.append("non ci siamo")
#-----------------------------------------------------#
        # Earnings Value > 40%
#-----------------------------------------------------#
        net_income_earnings_value = financials.loc['Net Income'].to_numpy()
        net_income_earnings_value = np.flip(net_income_earnings_value)
        tassi_crescita_earnings = []
        for i in range(0,len(net_income_earnings_value)):

            if (i > 0):
                # Calcolo i Tassi di Crescita
                element = net_income_earnings_value[i] / net_income_earnings_value[i - 1]
                element = (element - 1)*100
                element = round(element,3)
                tassi_crescita_earnings.append(element)
                element = None

            else:
                continue

        media_tassi_crescita_earn = st.mean(tassi_crescita_earnings)
        # Check Earnings Value --> 9° Multiplo
        if media_tassi_crescita_earn > 12 :
            verify.append("buono")
        else:
            verify.append("non ci siamo")
#-----------------------------------------------------#
        # FCF Value
#-----------------------------------------------------#
        url = 'https://finance.yahoo.com/quote/' + s + '/cash-flow'

        driver.get(url)
        time.sleep(10)
        righe = []
        indice = []

        tabella = driver.find_element(By.CLASS_NAME, "W\(100\%\).Whs\(nw\).Ovx\(a\).BdT.Bdtc\(\$seperatorColor\)")

        titoli_tabella_1 = tabella.find_element(By.CLASS_NAME, "D\(tbhg\)")
        titoli_span = titoli_tabella_1.find_elements(By.TAG_NAME, "span")
        for m in titoli_span:
            indice.append(m.text)

        indici_tabella_1 = tabella.find_elements(By.CLASS_NAME, "D\(tbr\).fi-row.Bgc\(\$hoverBgColor\)\:h")
        for i in indici_tabella_1:
            indici_tabella_2 = i.find_elements(By.TAG_NAME, "div")
            array = []
            for j in indici_tabella_2:
                span = j.find_elements(By.TAG_NAME, "span")
                for k in span:
                    number = (k.text).replace(",", "")
                    array.append(number)
            array = np.delete(array, 0)
            righe.append(array)

        tabelle = pd.DataFrame(righe, columns = indice)
        free_cash_flow_indice = tabelle[tabelle["Breakdown"]=="Free Cash Flow"].index.values
        free_cash_flow = tabelle.iloc[free_cash_flow_indice].to_numpy()
        # Free Cash Flow
        free_cash_flow = np.delete(free_cash_flow, 0)
        free_cash_flow = free_cash_flow.astype(np.float)
        free_cash_flow = [element * 1000 for element in free_cash_flow]
        lenght_fcf = len(free_cash_flow)
        if lenght_fcf > 4:
            free_cash_flow = np.delete(free_cash_flow, 0)


        tassi_crescita_fcf = []
        for i in range(0,len(free_cash_flow)):

            if (i > 0):
                # Calcolo i Tassi di Crescita
                element = free_cash_flow[i] / free_cash_flow[i - 1]
                element = (element - 1)
                element = round(element,3)
                tassi_crescita_fcf.append(element)
                element = None

            else:
                continue

        media_tassi_crescita_fcf = st.mean(tassi_crescita_fcf)

        # Check FCF Value --> 10° Multiplo
        if media_tassi_crescita_fcf > 0.09 :
            verify.append("buono")
        else:
            verify.append("non ci siamo")
#-----------------------------------------------------#
        # ROE > P/B
#-----------------------------------------------------#
        net_income = net_income_earnings_value[-1]
        total_equity = balance.loc['Total Stockholder Equity'].to_numpy()
        total_equity = np.flip(total_equity)
        total_equity = total_equity[-1]
        roe = net_income / total_equity
        # Check ROE > P/BV --> 11° Multiplo
        try:
            if roe > price_to_book :
                verify.append("buono")
            else:
                verify.append("non ci siamo")
        except:
            verify.append("errore")
#-----------------------------------------------------#
        # ROIC > WACC
#-----------------------------------------------------#
        # Short Long Term Debt
        try:
            short_debt = balance.loc['Short Long Term Debt'].to_numpy()
            # Short Long Term Debt - Flippato
            short_debt = np.flip(short_debt)
            short_debt_1 = short_debt[-1]
            check = math.isnan(short_debt_1)
            if check == True:
                short_debt = short_debt[-2]
            else:
                short_debt = short_debt_1
        except:
            short_debt = 0
        # Total Assets
        total_assets = balance.loc['Total Assets'].to_numpy()
        total_assets = np.flip(total_assets)
        total_assets = total_assets[-1]
        # Current Liabilities
        current_liab = balance.loc['Total Current Liabilities'].to_numpy()
        current_liab = np.flip(current_liab)
        current_liab = current_liab[-1]
        # Account Payable
        accounts_payable = balance.loc['Accounts Payable'].to_numpy()
        accounts_payable = np.flip(accounts_payable)
        accounts_payable = accounts_payable[-1]
        # Income Before Tax or Pretax Income
        income_before_tax = financials.loc['Income Before Tax'].to_numpy()
        income_before_tax = st.mean(income_before_tax)
        # Interst Expense or Interest Expense Not Operating (Negativo)
        interest_expense = financials.loc['Interest Expense'].to_numpy()
        lenght_interest_expense = len(interest_expense)
        if lenght_interest_expense >= 3:
            if(interest_expense[0] is not None):
                check = math.isnan(interest_expense[0])
                if check == True:
                    interest_expense = np.delete(interest_expense, 0)
                    interest_expense = np.negative(interest_expense)
                    interest_expense = st.mean(interest_expense)
                else:
                    interest_expense = np.negative(interest_expense)
                    interest_expense = st.mean(interest_expense)
            else:
                interest_expense = 0
        # Income Tax Expense or Tax Provision
        income_tax_expense = financials.loc['Income Tax Expense'].to_numpy()
        income_tax_expense = st.mean(income_tax_expense)
        # Calcolo del CAPM (Ce) - percent
        capm = rfr + (beta*(rdm-rfr))
        # Calcolo del Debt (D)
        debt = short_debt + long_term_debt
        # Calcolo del Costo del debito (Cd) - percent
        if debt > 0:
            cost_debt = interest_expense / debt
        else:
            cost_debt = 0
        # Calcolo del Tax Rate (T) - percent
        tax_rate = income_tax_expense / income_before_tax
        # Calcolo WACC
        wacc = (capm * (market_cap/(market_cap + debt))) + (cost_debt * (1 - tax_rate) * (debt/(debt + market_cap)))
        roic = (net_income*(1-(income_tax_expense/income_before_tax)))/(total_assets-accounts_payable+current_liab-current_assets)
        # Check ROIC > WACC --> 12° Multiplo
        if roic > wacc :
            verify.append("buono")
        else:
            verify.append("non ci siamo")

        mylist = [verify]
        for row in mylist:
            ws.append(row)
        wb.save(filename)
        print('Nome Azienda: ' + name + ' completato')
        verify = None
        mylist = None
        print('mi prendo una pausa')
        time.sleep(60)

    except:
        print(name + " " + "errore")
        mylist = [verify]
        for row in mylist:
            ws.append(row)
        wb.save(filename)
        verify = None
        mylist = None
        print('mi prendo una pausa')
        time.sleep(60)

wb.save(filename)
wb.close()
driver.close()
driver.quit()
print("Ok, ho finito")
