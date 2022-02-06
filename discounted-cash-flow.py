from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
import pandas as pd
import numpy as np
import yfinance as yf
import statistics as st
import openpyxl
import requests
import time
import difflib
import math

# Carico File Excel Watchlist
filename = 'PATH/WatchList.xlsx'
wb = openpyxl.load_workbook(filename)
ws = wb['YNAME']
stocks = []
column = ws['A']
for x in range(0, len(column)):
    stocks.append(column[x].value)
print('Ho caricato le lista Stocks')

# Carico Aziende già analizzate
ws = wb['DCF']
stocks_dcf = []
column = ws['A']
for x in range(1, len(column)):
    stocks_dcf.append(column[x].value)
print('Ho caricato le lista delle Stocks già analizzate')


# Carico Performance Indice S&P500 ultimi 10 anni
table_sp500 = pd.read_html('https://www.macrotrends.net/2526/sp-500-historical-annual-returns')
table_sp500 = table_sp500[0]
table_sp500 = table_sp500.values
annual_change = []
for t in table_sp500:
    p = t[6]
    annual_change.append(p)
annual_change = [element.replace('%', '') for element in annual_change]
sp500 = np.flip(annual_change)
sp500 = sp500[-10:]
sp500 = sp500.astype(np.float)
sp500 = [element / 100 for element in sp500]
print("Carico Performance Indice S&P500 ultimi 10 anni")
# Calcolo il Rendimento di Mercato
rdm = st.mean(sp500)
# Rendimento Free Risk -> Tasso di Interesse a 10 anni
table_wgb = pd.read_html('http://www.worldgovernmentbonds.com/')
table_wgb = table_wgb[1]
wgb_country = table_wgb["Country"].values
# Tasso di crescita PIL - GDP - Growth Rate
gdp = [2.56,1.55,2.25,1.84,2.53,3.08,1.71,2.33,3.00,2.16,-3.49]
gdp = [element / 100 for element in gdp]
gdp = st.mean(gdp)


# Carico Settings Selenium
options = webdriver.ChromeOptions()
options.add_extension("PATH/ublock.crx")
options.add_extension("PATH/cookie.crx")
options.add_argument('--disable-gpu')
options.add_argument("--log-level=0")
options.add_argument("--log-level=1")
options.add_argument("--log-level=2")
options.add_argument("--log-level=3")
options.add_argument("--ignore-certificate-errors")
options.add_argument("--start-maximized")
driver = webdriver.Chrome(executable_path=r'PATH/chromedriver.exe', options=options)

for i in stocks:
    try:
        # Yahoo Finance Stock
        stock = yf.Ticker(i)
        info = stock.info
        financials = stock.financials
        balance = stock.balance_sheet
        cashflow = stock.cashflow
        # Nome Stock
        name = info['shortName']
        if name in stocks_dcf:
            continue
        # Country
        country = info['country']
        # Prezzo Stock
        price = info['regularMarketPrice']
        # Sector
        sector = info['sector']
        # Industry
        industry = info['industry']
        # Beta Stock --> Calcolo in caso parametro mancante
        beta = info['beta']
        # MarketCap
        market_cap = info['marketCap']
        # N. Shares
        n_share = info['sharesOutstanding']
        # Rendimento Free Risk -> Tasso di Interesse a 10 anni
        match_country = difflib.get_close_matches(country, wgb_country)
        match_result_wgb = match_country[0]
        index_wgb = table_wgb[table_wgb["Country"]==match_result_wgb].index.values
        row_wgb = table_wgb.iloc[index_wgb].to_numpy()
        rfr = row_wgb[0,3]
        rfr = rfr.replace('%', '')
        rfr = float(rfr)
        rfr = rfr / 100

        url = 'https://finance.yahoo.com/quote/' + i + '/cash-flow'

        driver.get(url)
        time.sleep(10)
        righe = []
        indice = []

        tabella = driver.find_element(By.CLASS_NAME, "W\(100\%\).Whs\(nw\).Ovx\(a\).BdT.Bdtc\(\$seperatorColor\)")

        titoli_tabella_1 = tabella.find_element(By.CLASS_NAME, "D\(tbhg\)")
        titoli_span = titoli_tabella_1.find_elements(By.TAG_NAME, "span")
        for m in titoli_span:
            indice.append(m.text)

        indici_tabella_1 = tabella.find_elements(By.CLASS_NAME, "D\(tbr\).fi-row.Bgc\(\$hoverBgColor\)\:h")
        for i in indici_tabella_1:
            indici_tabella_2 = i.find_elements(By.TAG_NAME, "div")
            array = []
            for j in indici_tabella_2:
                span = j.find_elements(By.TAG_NAME, "span")
                for k in span:
                    number = (k.text).replace(",", "")
                    array.append(number)
            array = np.delete(array, 0)
            righe.append(array)

        tabelle = pd.DataFrame(righe, columns = indice)
        free_cash_flow_indice = tabelle[tabelle["Breakdown"]=="Free Cash Flow"].index.values
        free_cash_flow = tabelle.iloc[free_cash_flow_indice].to_numpy()
        # Free Cash Flow
        free_cash_flow = np.delete(free_cash_flow, 0)
        free_cash_flow = free_cash_flow.astype(np.float)
        free_cash_flow = [element * 1000 for element in free_cash_flow]
        lenght_fcf = len(free_cash_flow)
        if lenght_fcf > 4:
            free_cash_flow = np.delete(free_cash_flow, 0)

        # Short Long Term Debt
        try:
            short_debt = balance.loc['Short Long Term Debt'].to_numpy()
            # Short Long Term Debt - Flippato
            short_debt = np.flip(short_debt)
            short_debt_1 = short_debt[-1]
            check = math.isnan(short_debt_1)
            if check == True:
                short_debt = short_debt[-2]
            else:
                short_debt = short_debt_1
        except:
            short_debt = 0
        # Long Term Debt
        try:
            long_term_debt = balance.loc['Long Term Debt'].to_numpy()
            # Long Term Debt - Flippato
            long_term_debt = np.flip(long_term_debt)
            long_term_debt_1 = long_term_debt[-1]
            check = math.isnan(long_term_debt_1)
            if check == True:
                long_term_debt = long_term_debt[-2]
            else:
                long_term_debt = long_term_debt_1
        except:
            long_term_debt = 0
        # Net Income degli ultimi 4 anni
        net_income = financials.loc['Net Income'].to_numpy()
        # Income Before Tax or Pretax Income
        income_before_tax = financials.loc['Income Before Tax'].to_numpy()
        income_before_tax = st.mean(income_before_tax)
        # Income Tax Expense or Tax Provision
        income_tax_expense = financials.loc['Income Tax Expense'].to_numpy()
        income_tax_expense = st.mean(income_tax_expense)
        # Interst Expense or Interest Expense Not Operating (Negativo)
        interest_expense = financials.loc['Interest Expense'].to_numpy()
        lenght_interest_expense = len(interest_expense)
        if lenght_interest_expense >= 3:
            if(interest_expense[0] is not None):
                check = math.isnan(interest_expense[0])
                if check == True:
                    interest_expense = np.delete(interest_expense, 0)
                    interest_expense = np.negative(interest_expense)
                    interest_expense = st.mean(interest_expense)
                else:
                    interest_expense = np.negative(interest_expense)
                    interest_expense = st.mean(interest_expense)
            else:
                interest_expense = 0

        # Flip array
        # Net Income degli ultimi 4 anni - Flippato
        net_income = np.flip(net_income)
        # Free Cash Flow - Flippato
        free_cash_flow = np.flip(free_cash_flow)


        tassi_crescita = []
        for i in range(0,len(net_income)):

            if (i > 0):
                # Calcolo i Tassi di Crescita
                element = net_income[i] / net_income[i - 1]
                element = (element - 1)*100
                element = round(element,3)
                tassi_crescita.append(element)
                element = None

            else:
                continue

        # Calcolo Media dei Tassi di crescita
        media_tassi_crescita = st.mean(tassi_crescita)
        # Calcolo il Dividendo Futuro Atteso
        net_income_futuro_atteso_1 = net_income[-1] * (1+(media_tassi_crescita/100))
        net_income_futuro_atteso_2 = net_income_futuro_atteso_1 * (1+(media_tassi_crescita/100))
        net_income_futuro_atteso_3 = net_income_futuro_atteso_2 * (1+(media_tassi_crescita/100))
        net_income_futuro_atteso_4 = net_income_futuro_atteso_3 * (1+(media_tassi_crescita/100))
        # Media Free Cash Flow / Net Income
        try:
            divide_fcd_net_income = np.divide(free_cash_flow, net_income)
        except:
            free_cash_flow = np.delete(free_cash_flow, -1)
            divide_fcd_net_income = np.divide(free_cash_flow, net_income)

        media_fcd_net_income = st.mean(divide_fcd_net_income)
        # Free Cash Flow Futuro Atteso
        free_cash_flow_futuro_atteso_1 = net_income_futuro_atteso_1 * media_fcd_net_income
        free_cash_flow_futuro_atteso_2 = net_income_futuro_atteso_2 * media_fcd_net_income
        free_cash_flow_futuro_atteso_3 = net_income_futuro_atteso_3 * media_fcd_net_income
        free_cash_flow_futuro_atteso_4 = net_income_futuro_atteso_4 * media_fcd_net_income
        # Calcolo del CAPM (Ce) - percent
        capm = rfr + (beta*(rdm-rfr))
        # Calcolo del Debt (D)
        debt = short_debt + long_term_debt
        # Calcolo del Costo del debito (Cd) - percent
        if debt > 0:
            cost_debt = interest_expense / debt
        else:
            cost_debt = 0
        # Calcolo del Tax Rate (T) - percent
        tax_rate = income_tax_expense / income_before_tax
        # Calcolo WACC
        wacc = (capm * (market_cap/(market_cap + debt))) + (cost_debt * (1 - tax_rate) * (debt/(debt + market_cap)))
        # WACC Capitalizzato
        wacc_capitalizzato_1 = 1 + wacc
        wacc_capitalizzato_2 = pow((1 + wacc),2)
        wacc_capitalizzato_3 = pow((1 + wacc),3)
        wacc_capitalizzato_4 = pow((1 + wacc),4)
        # Free Cash Flow Attualizzato
        free_cash_flow_attualizzato_1 = free_cash_flow_futuro_atteso_1 / wacc_capitalizzato_1
        free_cash_flow_attualizzato_2 = free_cash_flow_futuro_atteso_2 / wacc_capitalizzato_2
        free_cash_flow_attualizzato_3 = free_cash_flow_futuro_atteso_3 / wacc_capitalizzato_3
        free_cash_flow_attualizzato_4 = free_cash_flow_futuro_atteso_4 / wacc_capitalizzato_4
        # Terminal Value
        terminal_value = (free_cash_flow_futuro_atteso_4 * (1 + gdp)) / (wacc - gdp)
        terminal_value = terminal_value / wacc_capitalizzato_4
        # Valore Attualizzato
        valore_attualizzato = free_cash_flow_attualizzato_1 + free_cash_flow_attualizzato_2 + free_cash_flow_attualizzato_3 + free_cash_flow_attualizzato_4 + terminal_value
        # Calcolo del Price Fair Value
        price_fair_value = valore_attualizzato / n_share
        price_fair_value = round(price_fair_value,3)

        price = str(price)
        price_fair_value = str(price_fair_value)

        mylist = [
            [name,price_fair_value]
            ]
        for row in mylist:
            ws.append(row)
        wb.save(filename)
        print('Il Price Fair Value di ' + name + ' è: ' + price_fair_value)
        print("")
        mylist.clear()
        print('mi prendo una pausa')
        time.sleep(60)
    except:
        print(name + " " + "errore")
        print("")
        mylist = [
            [name,"errore"]
            ]
        for row in mylist:
            ws.append(row)
        wb.save(filename)
        mylist.clear()
        print('mi prendo una pausa')
        time.sleep(60)


wb.save(filename)
wb.close()
driver.close()
driver.quit()
print("Ok, ho finito")
